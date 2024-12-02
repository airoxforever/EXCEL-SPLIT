import pandas as pd
from utils.logging_config import setup_logging
from config import SUPPORTED_LANGUAGES, SOURCE_LANGUAGE
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import tempfile
import os
from pathlib import Path
from datetime import datetime
import re
import shutil
from openpyxl.styles import Color
from copy import copy
from openpyxl import Workbook

# Set up logging to a file in a logs subdirectory
log_dir = Path("logs")
log_dir.mkdir(exist_ok=True)
logger = setup_logging("excel_processor")

class ExcelProcessor:
    def __init__(self, file):
        self.file = file
        self.df = None
        self.wb = None
        
    def clean_formatted_text(self, text_parts):
        """
        Clean and properly format text with tags, handling both bold and red text
        """
        result = []
        current_formatted_text = []
        
        # Check if the entire cell is red (all parts have red formatting)
        all_red = True
        for part in text_parts:
            if isinstance(part, TextBlock):
                if not (part.font and part.font.color and part.font.color.rgb == "FFFF0000"):
                    all_red = False
                    break
        
        # If entire cell is red, wrap the whole content
        if all_red:
            full_text = ''.join(str(part.text) if isinstance(part, TextBlock) else str(part) for part in text_parts)
            return f"<cr>{full_text}</cr>"
        
        # Otherwise, process parts individually
        for part in text_parts:
            if isinstance(part, TextBlock):
                text = str(part.text)
                is_bold = part.font and part.font.b
                is_red = part.font and part.font.color and part.font.color.rgb == "FFFF0000"  # Red color in RGB
                
                # If text is just whitespace, add it directly
                if text.isspace():
                    result.append(text)
                    continue
                    
                # Split text but preserve spaces using regex
                words = re.split(r'(\s+)', text)
                
                for word in words:
                    if word.isspace():  # Handle spacing
                        if current_formatted_text:
                            current_formatted_text.append(word)
                        else:
                            result.append(word)
                    elif word:  # Handle non-empty words
                        if is_bold and is_red:
                            result.append(f"<cfr>{word}</cfr>")  # Combined bold and red
                        elif is_bold:
                            result.append(f"<cf>{word}</cf>")    # Bold only
                        elif is_red:
                            result.append(f"<cr>{word}</cr>")    # Red only
                        else:
                            result.append(word)
            else:
                result.append(str(part))
        
        # Clean up multiple spaces and ensure proper spacing around tags
        text = ''.join(result)
        
        # Fix spacing around tags with punctuation and sentence awareness
        for tag in ['<cf>', '<cr>', '<cfr>']:
            # Add space before tag if preceded by non-space, non-punctuation, and not at start
            text = re.sub(r'(?<!^)(?<![\s\.,!?\)])(' + re.escape(tag) + ')', r' \1', text)
        
        for tag in ['</cf>', '</cr>', '</cfr>']:
            # Add space after tag if followed by non-space, non-punctuation
            text = re.sub(r'(' + re.escape(tag) + r')(?![\s\.,!?\)])(?!$)', r'\1 ', text)
        
        return text

    def read_excel(self, skip_first_row=True):
        """
        Read Excel file and immediately process formatting into tags
        """
        try:
            # Create temporary file for openpyxl
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                # Handle both file objects and file paths
                if isinstance(self.file, (str, Path)):
                    # If it's a path, copy the file
                    shutil.copy(str(self.file), tmp.name)
                else:
                    # If it's a file object, write its contents
                    self.file.seek(0)
                    tmp.write(self.file.read())
                tmp_path = tmp.name
            
            # Load with openpyxl first to get formatting
            self.wb = load_workbook(tmp_path, data_only=True, rich_text=True)
            ws = self.wb.active
            
            # Process formatting immediately
            formatted_cells = {}
            logger.info("Processing cell formatting...")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2 if skip_first_row else 1)):
                for cell in row:
                    if isinstance(cell.value, CellRichText):
                        # Use the new clean_formatted_text method
                        formatted_text = self.clean_formatted_text(cell.value)
                        if any(tag in formatted_text for tag in ['<cf>', '<cr>', '<cfr>']):
                            logger.debug(f"Found rich text in cell {cell.coordinate}: {formatted_text}")
                            formatted_cells[cell.coordinate] = formatted_text
                    elif cell.font:
                        # Check for whole-cell formatting
                        text = str(cell.value)
                        if text.strip():  # Only process non-empty cells
                            if cell.font.color and cell.font.color.rgb == "FFFF0000":
                                # Whole cell is red
                                formatted_cells[cell.coordinate] = f"<cr>{text}</cr>"
                                logger.debug(f"Found red cell {cell.coordinate}: {text}")
                            elif cell.font.b:
                                # Whole cell is bold
                                formatted_cells[cell.coordinate] = f"<cf>{text}</cf>"
                                logger.debug(f"Found bold cell {cell.coordinate}: {text}")
            
            # Now read with pandas
            self.df = pd.read_excel(tmp_path, header=1 if skip_first_row else 0)
            
            # Apply formatted text to DataFrame
            for coord, text in formatted_cells.items():
                # Extract column letter and row number correctly
                col_letter = ''.join(c for c in coord if c.isalpha())
                row_num = int(''.join(c for c in coord if c.isdigit()))
                
                # Convert to zero-based index
                col_idx = column_index_from_string(col_letter) - 1
                row_idx = row_num - (2 if skip_first_row else 1) - 1
                
                # Apply the formatted text if indices are valid
                if row_idx >= 0 and row_idx < len(self.df) and col_idx >= 0 and col_idx < len(self.df.columns):
                    logger.debug(f"Applying formatted text at row {row_idx}, col {col_idx}: {text}")
                    self.df.iloc[row_idx, col_idx] = text
            
            # Save debug Excel file with tags
            debug_output = Path("debug_output")
            debug_output.mkdir(exist_ok=True)
            debug_file = debug_output / f"debug_tagged_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.df.to_excel(debug_file, index=False)
            logger.info(f"Saved debug Excel file with tags: {debug_file}")
            
            # Create verification CSV with original and tagged text
            verification_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2 if skip_first_row else 1)):
                for cell in row:
                    if cell.coordinate in formatted_cells:
                        verification_data.append({
                            'Cell': cell.coordinate,
                            'Original': str(cell.value),
                            'Tagged': formatted_cells[cell.coordinate]
                        })
            
            if verification_data:
                verification_df = pd.DataFrame(verification_data)
                csv_file = debug_output / f"tag_verification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                verification_df.to_csv(csv_file, sep=';', index=False, encoding='utf-8-sig')
                logger.info(f"Saved tag verification CSV: {csv_file}")
            
            return self.df
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
            raise
        finally:
            if 'tmp_path' in locals():
                try:
                    os.unlink(tmp_path)
                except:
                    pass

    def get_cell_formatting(self, row_idx, col_name):
        """Get formatting information for a specific cell"""
        if self.wb is None:
            return None
            
        ws = self.wb.active
        # Adjust row index based on header configuration
        excel_row = row_idx + 3  # Adjust for 0-based index and header rows
        
        # Find column letter from column name
        col_letter = None
        for idx, cell in enumerate(ws[1], 1):  # Assuming headers are in row 1
            if cell.value == col_name:
                col_letter = cell.column_letter
                break
                
        if col_letter:
            cell = ws[f"{col_letter}{excel_row}"]
            logger.info(f"Checking formatting for cell {col_letter}{excel_row}")
            
            # Handle rich text
            if isinstance(cell.value, CellRichText):
                logger.info(f"Found rich text formatting")
                return cell.value
            elif cell.font and (cell.font.bold or cell.font.color):
                # Create rich text for cells with direct formatting
                rich_text = CellRichText()
                
                # Convert RGB color to Color object if needed
                color = None
                if cell.font.color:
                    if isinstance(cell.font.color, Color):
                        color = cell.font.color
                    else:
                        # Create a new Color object from the RGB value
                        rgb_value = cell.font.color.rgb if hasattr(cell.font.color, 'rgb') else None
                        if rgb_value:
                            color = Color(rgb=rgb_value)
                
                inline_font = InlineFont(
                    b=cell.font.bold,
                    color=color
                )
                rich_text.append(TextBlock(inline_font, str(cell.value)))
                return rich_text
            
        return None

    def detect_languages(self):
        """Detect language columns based on supported language codes"""
        if self.df is None:
            self.read_excel()
        
        detected_languages = []
        for col in self.df.columns:
            # Check if column name contains a language code
            for lang_code in SUPPORTED_LANGUAGES.keys():
                if lang_code in str(col):  # Convert to string to handle any non-string column names
                    detected_languages.append(lang_code)
                    logger.info(f"Detected language column: {lang_code} in '{col}'")
                    break
        
        if not detected_languages:
            logger.warning("No supported language columns detected")
        
        return detected_languages

    def validate_source_language(self):
        """Validate that source language (English GB) exists in the Excel file"""
        languages = self.detect_languages()
        if SOURCE_LANGUAGE not in languages:
            logger.error(f"Source language {SOURCE_LANGUAGE} not found in Excel file")
            raise ValueError(f"Source language {SOURCE_LANGUAGE} not found in Excel file")
        return True

    def get_available_languages(self):
        """Get list of available languages in the Excel file"""
        return self.detect_languages()

    def preserve_workbook_format(self, original_file):
        """
        Load and preserve the original workbook format including column widths
        """
        temp_file = None
        try:
            if isinstance(original_file, (str, Path)):
                self.wb = load_workbook(original_file)
            else:
                # Create temporary file for file object
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                original_file.seek(0)
                temp_file.write(original_file.read())
                temp_file.close()  # Close the temp file before loading
                self.wb = load_workbook(temp_file.name)
                # Close workbook after loading
                self.wb.close()
                # Reopen workbook
                self.wb = load_workbook(temp_file.name)
            
            self.ws = self.wb.active
            # Store column dimensions
            self.column_dimensions = {col: self.ws.column_dimensions[col].width 
                                   for col in self.ws.column_dimensions}
            logger.info("Preserved original workbook formatting")
            return True
        except Exception as e:
            logger.error(f"Error preserving workbook format: {str(e)}", exc_info=True)
            return False
        finally:
            # Clean up temp file if it exists
            if temp_file:
                try:
                    if self.wb:
                        self.wb.close()
                    os.unlink(temp_file.name)
                except Exception as e:
                    logger.warning(f"Could not delete temporary file {temp_file.name}: {e}")

    def apply_translations_to_workbook(self, translations_dict):
        """
        Apply translations to the original workbook while preserving formatting
        """
        try:
            if not self.wb:
                raise ValueError("No workbook loaded. Call preserve_workbook_format first.")

            ws = self.wb.active
            
            # Find header row (usually row 1)
            header_row = 1
            
            # Create column mapping
            column_mapping = {}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                if cell.value:
                    for lang_code in translations_dict.keys():
                        if lang_code in str(cell.value):
                            column_mapping[lang_code] = col
            
            # Find start row for translations (skip header)
            start_row = header_row + 1
            
            # Apply translations while preserving formatting
            for lang_code, translation_df in translations_dict.items():
                if lang_code in column_mapping:
                    col_idx = column_mapping[lang_code]
                    
                    for idx, translation in enumerate(translation_df['Target'].values, start=start_row):
                        # Only update if there's actual content
                        if pd.notna(translation):
                            cell = ws.cell(row=idx, column=col_idx)
                            cell.value = translation
            
            # Ensure column widths are preserved
            for col, width in self.column_dimensions.items():
                if width is not None:
                    ws.column_dimensions[col].width = width
            
            logger.info("Successfully applied translations while preserving formatting")
            return True
            
        except Exception as e:
            logger.error(f"Error applying translations: {str(e)}", exc_info=True)
            return False

    def save_workbook(self, output_path):
        """
        Save the workbook while preserving all formatting
        """
        try:
            if not self.wb:
                raise ValueError("No workbook loaded")
            
            self.wb.save(output_path)
            logger.info(f"Saved formatted workbook to {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}", exc_info=True)
            return False

    def detect_header_row(self):
        """
        Scan the Excel file to find the row containing language codes
        Returns the header row number (1-based) or None if not found
        """
        try:
            ws = self.wb.active
            max_rows_to_scan = min(10, ws.max_row)  # Scan first 10 rows or less
            
            for row_num in range(1, max_rows_to_scan + 1):
                language_found = False
                for cell in ws[row_num]:
                    cell_value = str(cell.value or '').lower()
                    # Look for language codes in the cell value
                    for lang_code in SUPPORTED_LANGUAGES.keys():
                        if lang_code.lower() in cell_value:
                            language_found = True
                            break
                    if language_found:
                        return row_num
            return None
        except Exception as e:
            logger.error(f"Error detecting header row: {e}")
            return None

    def get_column_info(self):
        """
        Get detailed information about language columns
        Returns a dict with column information including indexes and headers
        """
        try:
            if not self.wb:
                raise ValueError("Workbook not loaded")

            ws = self.wb.active
            header_row = self.detect_header_row()
            if not header_row:
                return None

            columns_info = {
                'header_row': header_row,
                'columns': {},
                'source_column': None
            }

            # Scan header row for language columns
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                cell_value = str(cell.value or '').lower()
                
                # Check for language codes
                for lang_code, lang_name in SUPPORTED_LANGUAGES.items():
                    if lang_code.lower() in cell_value:
                        col_info = {
                            'index': col,
                            'header': cell.value,
                            'letter': get_column_letter(col),
                            'width': ws.column_dimensions[get_column_letter(col)].width,
                            'language_code': lang_code,
                            'language_name': lang_name
                        }
                        columns_info['columns'][lang_code] = col_info
                        
                        # Mark as source if it's English GB
                        if lang_code == SOURCE_LANGUAGE:
                            columns_info['source_column'] = col_info

            return columns_info
        except Exception as e:
            logger.error(f"Error getting column info: {e}")
            return None

    def copy_cell_format(self, source_cell, target_cell):
        """
        Copy all formatting from source cell to target cell with proper handling
        """
        if source_cell.has_style:
            try:
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
            except Exception as e:
                logger.warning(f"Could not copy some cell formatting: {e}")

    def create_bilingual_file(self, source_lang, target_lang):
        """
        Create a bilingual Excel file by copying and modifying the source workbook
        """
        try:
            if not self.wb:
                raise ValueError("Workbook not loaded")

            # Get column information
            col_info = self.get_column_info()
            if not col_info:
                raise ValueError("Could not detect language columns")

            source_col = col_info['columns'].get(source_lang)
            target_col = col_info['columns'].get(target_lang)
            
            if not source_col or not target_col:
                raise ValueError(f"Could not find columns for {source_lang} and/or {target_lang}")

            # Create a copy of the source workbook
            new_wb = Workbook()
            new_ws = new_wb.active

            # Copy worksheet properties from source
            ws = self.wb.active
            new_ws.sheet_format = copy(ws.sheet_format)
            new_ws.sheet_properties = copy(ws.sheet_properties)
            
            # Copy column widths
            new_ws.column_dimensions['A'].width = source_col['width'] if source_col['width'] else 10.0
            new_ws.column_dimensions['B'].width = target_col['width'] if target_col['width'] else 10.0

            # Set headers with original styles if available
            source_header = ws.cell(row=col_info['header_row'], column=source_col['index'])
            target_header = ws.cell(row=col_info['header_row'], column=target_col['index'])
            
            header_a = new_ws['A1']
            header_b = new_ws['B1']
            
            # Copy header values and styles
            header_a.value = source_col['header']
            header_b.value = target_col['header']
            
            self.copy_cell_format(source_header, header_a)
            self.copy_cell_format(target_header, header_b)

            # Copy content and formatting
            row_offset = col_info['header_row']
            new_row = 2  # Start after header

            for row in range(row_offset + 1, ws.max_row + 1):
                source_cell = ws.cell(row=row, column=source_col['index'])
                target_cell = ws.cell(row=row, column=target_col['index'])
                
                if source_cell.value is None and target_cell.value is None:
                    continue  # Skip completely empty rows
                
                # Copy values and formatting
                new_source = new_ws.cell(row=new_row, column=1)
                new_target = new_ws.cell(row=new_row, column=2)
                
                # Preserve original values without conversion when possible
                new_source.value = source_cell.value
                new_target.value = target_cell.value
                
                # Copy all formatting
                self.copy_cell_format(source_cell, new_source)
                self.copy_cell_format(target_cell, new_target)
                
                new_row += 1

            # Copy workbook properties
            if hasattr(self.wb, 'properties'):
                new_wb.properties = copy(self.wb.properties)
            
            # Ensure core properties are set
            new_wb.properties.creator = "Excel File Splitter"
            new_wb.properties.created = datetime.now()
            new_wb.properties.modified = datetime.now()

            return new_wb
        except Exception as e:
            logger.error(f"Error creating bilingual file: {e}")
            raise

    def create_header_style(self):
        """Create a proper style for headers"""
        from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
        
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True, size=11)
        header_style.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        header_style.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        
        return header_style