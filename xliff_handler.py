from lxml import etree
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from utils.logging_config import setup_logging
from excel_handler import ExcelProcessor
from config import SUPPORTED_LANGUAGES
import re
import shutil
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.colors import Color
from itertools import zip_longest
from utils.sentence_splitter import SentenceSplitter, try_split_segments
import streamlit as st
import os

logger = setup_logging("xliff_handler")

def get_column_letter_index(column_letter):
    """Convert column letter to column index"""
    return column_index_from_string(column_letter)

class XliffHandler:
    def __init__(self):
        self.nsmap = {
            None: "urn:oasis:names:tc:xliff:document:1.2",
            "xsi": "http://www.w3.org/2001/XMLSchema-instance"
        }
        self.processor = None
        self.sentence_splitter = SentenceSplitter(
            min_segment_length=5,
            max_unsplit_length=70
        )
        self.processing_stats = {}  # Store stats per language

    def format_rich_text(self, cell_content, element):
        """Convert rich text to plain text with custom tags for bold text"""
        try:
            if isinstance(cell_content, CellRichText):
                logger.debug(f"Processing rich text: {cell_content}")
                result = []
                
                for part in cell_content:
                    if isinstance(part, TextBlock):
                        text = part.text
                        font = part.font
                        
                        if font and font.b:
                            logger.debug(f"Found bold text: {text}")
                            result.append(f"<cf>{text}</cf>")
                        else:
                            result.append(text)
                    else:
                        result.append(str(part))
                
                final_text = "".join(result)
                logger.debug(f"Final formatted text: {final_text}")
                element.text = final_text
                return True
            else:
                logger.debug("Not rich text, using as-is")
                element.text = str(cell_content)
                return False
                
        except Exception as e:
            logger.error(f"Error in format_rich_text: {e}", exc_info=True)
            element.text = str(cell_content)
            return False

    def create_xliff(self, df, target_lang, excel_path):
        """Create XLIFF document structure"""
        xliff = etree.Element("xliff", 
            version="1.2",
            nsmap=self.nsmap,
            attrib={
                "{http://www.w3.org/2001/XMLSchema-instance}schemaLocation": 
                    "urn:oasis:names:tc:xliff:document:1.2 xliff-core-1.2-strict.xsd"
            }
        )
        
        # Special handling for Norwegian
        target_lang_code = target_lang.replace('_', '-')
        if target_lang_code.lower() == 'no-no':
            target_lang_code = 'nb-NO'  # Use correct Norwegian Bokmål code
        
        file_elem = etree.SubElement(xliff, "file",
            attrib={
                "original": str(excel_path),
                "datatype": "plaintext",
                "source-language": "en-US",
                "target-language": target_lang_code
            }
        )
        
        body = etree.SubElement(file_elem, "body")
        
        source_col = next(col for col in df.columns if "en_GB" in col)
        target_col = next(col for col in df.columns if target_lang in col)
        
        # Update stats structure to include more comment details
        self.processing_stats[target_lang] = {
            'total_segments': 0,
            'split_attempts': 0,
            'successful_splits': 0,
            'segments_with_comments': 0,
            'comment_details': []  # Store details about segments with comments
        }
        
        for idx, row in df.iterrows():
            source_text = row[source_col]
            target_text = row[target_col] if pd.notna(row[target_col]) else ""
            
            self.processing_stats[target_lang]['total_segments'] += 1
            
            if pd.isna(source_text) or str(source_text).lower() == 'nan':
                continue
            
            # Get comment if exists
            comment_col = st.session_state.comment_column
            comment = str(row.get(comment_col, '')).strip()
            
            if comment:
                self.processing_stats[target_lang]['segments_with_comments'] += 1
                self.processing_stats[target_lang]['comment_details'].append({
                    'segment': idx + 1,
                    'source_text': str(source_text)[:50] + '...' if len(str(source_text)) > 50 else str(source_text),
                    'comment': comment[:50] + '...' if len(comment) > 50 else comment
                })
            
            # Check if splitting is enabled
            if st.session_state.enable_splitting:
                # Try to split both texts
                if len(str(source_text)) > self.sentence_splitter.max_unsplit_length:
                    self.processing_stats[target_lang]['split_attempts'] += 1
                
                source_segments, target_segments = try_split_segments(
                    str(source_text),
                    str(target_text),
                    self.sentence_splitter
                )
                
                if len(source_segments) > 1:
                    self.processing_stats[target_lang]['successful_splits'] += 1
            else:
                # If splitting is disabled, treat each row as a single segment
                source_segments = [str(source_text)]
                target_segments = [str(target_text)]
            
            # Create trans-unit for each segment
            for sent_idx, (src_sent, tgt_sent) in enumerate(zip(source_segments, target_segments)):
                trans_unit = etree.SubElement(body, "trans-unit",
                    attrib={
                        "id": f"_msg{idx}_{sent_idx}",
                        "datatype": "plaintext"
                    }
                )
                
                source = etree.SubElement(trans_unit, "source",
                    attrib={
                        "{http://www.w3.org/XML/1998/namespace}space": "preserve"
                    }
                )
                
                # Process formatting
                if self.processor:
                    rich_text = self.processor.get_cell_formatting(idx, source_col)
                    if rich_text:
                        self.format_rich_text(rich_text, source)
                    else:
                        source.text = str(src_sent)
                else:
                    source.text = str(src_sent)
                
                target = etree.SubElement(trans_unit, "target",
                    attrib={
                        "state": "needs-review-translation",
                        "{http://www.w3.org/XML/1998/namespace}space": "preserve"
                    }
                )
                target.text = str(tgt_sent)
                
                # Add comment if exists
                if comment:
                    note = etree.SubElement(trans_unit, "note", 
                        attrib={
                            "from": "reviewer",
                            "annotates": "general",
                            "priority": "1"
                        }
                    )
                    note.text = comment
        
        return xliff

    def convert_to_xliff(self, df, target_languages, excel_path=None):
        """Convert Excel content to XLIFF files - one file per language"""
        if excel_path:
            excel_path = Path(excel_path).resolve()  # Get absolute path
            # Create main output directory if it doesn't exist
            main_output = Path("output")
            main_output.mkdir(exist_ok=True)
            
            # Create timestamped subfolder with shortened filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            shortened_name = excel_path.stem[:10]  # Take first 10 chars of filename
            output_dir = main_output / f"{shortened_name}_{timestamp}"
            output_dir.mkdir(exist_ok=True)
            
            logger.info(f"Creating XLIFF files in: {output_dir}")
        else:
            excel_path = Path("unknown_source.xlsx")
            output_dir = Path("output/unknown_source")
            output_dir.mkdir(parents=True, exist_ok=True)
        
        for target_lang in target_languages:
            if target_lang == "en_GB" or target_lang == "en_US":
                continue
                
            logger.info(f"Processing {target_lang}")
            
            try:
                xliff = self.create_xliff(df, target_lang, excel_path)
                output_filename = f"translation_{target_lang}"
                if target_lang == 'no_NO':
                    output_filename = "translation_nb_NO"
                
                output_file = output_dir / f"{output_filename}.xlf"
                
                tree = etree.ElementTree(xliff)
                tree.write(
                    str(output_file),
                    pretty_print=True,
                    xml_declaration=True,
                    encoding="utf-8"
                )
                
                logger.info(f"Created XLIFF file: {output_file}")
            except Exception as e:
                logger.error(f"Error processing language {target_lang}: {str(e)}")
                raise
        
        return True

    def xliff_to_excel(self, original_excel_path, xliff_folder_path, source_col="E", trans_start_pos="F3"):
        """
        Update Excel file with translations from XLIFF files in language-specific folders
        """
        logger.info(f"Processing XLIFF files from: {xliff_folder_path}")
        logger.info(f"Using source column: {source_col}")
        logger.info(f"Translations start at: {trans_start_pos}")
        
        # Create debug directory
        debug_dir = Path("debug_output")
        debug_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Parse translation start position
        trans_col = ''.join(filter(str.isalpha, trans_start_pos))
        trans_row = int(''.join(filter(str.isdigit, trans_start_pos)))
        
        # Create a copy of the original file
        original_path = Path(original_excel_path)
        reviewed_path = original_path.parent / f"{original_path.stem}_reviewed{original_path.suffix}"
        
        # If reviewed file exists, remove it first
        if reviewed_path.exists():
            logger.info(f"Removing existing reviewed file: {reviewed_path}")
            reviewed_path.unlink()
        
        # Create exact copy using shutil (preserves all metadata and formatting)
        logger.info(f"Creating new reviewed file: {reviewed_path}")
        shutil.copy2(original_path, reviewed_path)
        
        # Load the reviewed file for updating
        wb = load_workbook(reviewed_path, rich_text=True)
        ws = wb.active
        
        # Process XLIFF files
        processed_languages = []
        
        for lang_code in SUPPORTED_LANGUAGES:
            if lang_code in ['en_GB', 'en_US']:
                continue
                
            # Convert underscore to hyphen for folder lookup
            hyphen_code = lang_code.replace('_', '-').lower()
            if hyphen_code == 'no_no':
                hyphen_code = 'nb-no'
                
            lang_folder = Path(xliff_folder_path) / hyphen_code
            
            if not lang_folder.exists():
                logger.warning(f"No folder found for language {lang_code}")
                continue
                
            xliff_files = list(lang_folder.glob('translation_*.xlf'))
            xliff_files = [f for f in xliff_files if not f.name.endswith('.sdlxliff')]
            
            if not xliff_files:
                logger.warning(f"No XLIFF file found in {lang_folder}")
                continue
                
            xliff_file = xliff_files[0]
            
            try:
                # Get column index for this language
                lang_col = None
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=trans_row-1, column=col)
                    if lang_code in str(cell.value or ''):
                        lang_col = col
                        break
                
                if not lang_col:
                    logger.warning(f"No column found for language {lang_code}")
                    continue
                
                # Parse XLIFF and store translations
                tree = etree.parse(str(xliff_file))
                root = tree.getroot()
                
                # Dictionary to store segments by row
                row_segments = {}
                
                # First pass: collect all segments and group by row
                for trans_unit in root.findall('.//{urn:oasis:names:tc:xliff:document:1.2}trans-unit'):
                    try:
                        # Parse segment ID (e.g., "_msg5_1" -> row=5, segment=1)
                        seg_id = trans_unit.get('id')
                        row_match = re.match(r'_msg(\d+)_(\d+)', seg_id)
                        if not row_match:
                            logger.warning(f"Invalid segment ID format: {seg_id}")
                            continue
                        
                        row_num = int(row_match.group(1))
                        seg_num = int(row_match.group(2))
                        
                        target = trans_unit.find('.//{urn:oasis:names:tc:xliff:document:1.2}target')
                        target_text = target.text if target is not None and target.text else ""
                        
                        # Store segment with its order number
                        if row_num not in row_segments:
                            row_segments[row_num] = []
                        row_segments[row_num].append((seg_num, target_text))
                        
                    except ValueError as ve:
                        logger.error(f"Error processing segment: {ve}")
                        continue
                
                # Second pass: reconstruct and update cells
                for row_num, segments in row_segments.items():
                    try:
                        # Sort segments by their order number
                        segments.sort(key=lambda x: x[0])
                        
                        # Join segments with proper spacing
                        reconstructed_text = ""
                        for i, (_, text) in enumerate(segments):
                            if i > 0 and text and not text[0] in '.!?。！？':  # Don't add space before punctuation
                                reconstructed_text += " "
                            reconstructed_text += text
                        
                        # Calculate Excel row (row_num + trans_row)
                        excel_row = row_num + trans_row
                        
                        # Update cell with reconstructed text
                        cell = ws.cell(row=excel_row, column=lang_col)
                        
                        # Handle rich text formatting if present
                        if any(tag in reconstructed_text for tag in ['<cf>', '<cr>', '<cfr>']):
                            rich_text = self._create_rich_text(reconstructed_text)
                            cell.value = rich_text
                        else:
                            cell.value = reconstructed_text
                        
                    except Exception as e:
                        logger.error(f"Error updating row {row_num}: {str(e)}")
                        continue
                
                processed_languages.append(lang_code)
                
            except Exception as e:
                logger.error(f"Error processing {xliff_file}: {e}", exc_info=True)
                continue
        
        # Save the final reviewed file
        wb.save(reviewed_path)
        wb.close()
        
        logger.info(f"Saved updated translations to: {reviewed_path}")
        logger.info(f"Successfully processed languages: {', '.join(processed_languages)}")
        
        return reviewed_path

    def _create_rich_text(self, text):
        """Helper method to create rich text from tagged content"""
        rich_text = CellRichText()
        parts = re.split(r'(</?cf>|</?cr>|</?cfr>)', text)
        
        current_format = None
        for part in parts:
            if part in ['<cf>', '<cr>', '<cfr>']:
                current_format = part[1:-1]
            elif part in ['</cf>', '</cr>', '</cfr>']:
                current_format = None
            elif part:  # Text content
                color = Color(rgb='FFFF0000') if current_format in ['cr', 'cfr'] else None
                font = InlineFont(
                    b=current_format in ['cf', 'cfr'],
                    color=color
                )
                rich_text.append(TextBlock(font, part))
        
        return rich_text

    def update_settings(self, min_segment_length, max_unsplit_length):
        """Update sentence splitter settings"""
        self.sentence_splitter = SentenceSplitter(
            min_segment_length=min_segment_length,
            max_unsplit_length=max_unsplit_length
        )